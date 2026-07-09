"""
tracker.py v5.4 — Guarda resultados en Excel automáticamente.
Requiere: pip install openpyxl
"""
from __future__ import annotations
import os
from datetime import datetime
from typing import Optional

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

TRACKER_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "trading_tracker.xlsx"
)


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)  # type: ignore[return-value]


def _border() -> "Border":
    s = Side(style="thin")  # type: ignore[call-arg]
    return Border(left=s, right=s, top=s, bottom=s)  # type: ignore[return-value]


def _font(bold: bool = False, color: str = "000000", size: int = 10) -> "Font":
    return Font(bold=bold, color=color, size=size, name="Arial")  # type: ignore[return-value]


def _verdict(sharpe: float, win_rate: float, trades: int) -> str:
    if trades < 10:
        return "Pocos trades (<10)"
    if sharpe >= 2 and win_rate >= 55:
        return "EXCELENTE"
    if sharpe >= 1 and win_rate >= 45:
        return "BUENO"
    if sharpe >= 0:
        return "DEBIL"
    return "NEGATIVO"


def save_run(
    metrics: dict,
    config_name: str,
    version: str = "v4",
    notes: str = "",
    strategy_mode: str = "all",
    extra: Optional[dict] = None,
) -> bool:
    """
    Agrega una fila al Excel tracker con los resultados del run.
    Retorna True si se guardó correctamente.
    """
    if not OPENPYXL_OK:
        print("   ⚠️  openpyxl no disponible. Instalar: pip install openpyxl")
        return False

    if not os.path.exists(TRACKER_PATH):
        print(f"   ⚠️  Tracker no encontrado: {TRACKER_PATH}")
        print("      Descarga trading_tracker.xlsx y colócalo en la carpeta del proyecto.")
        return False

    try:
        wb = load_workbook(TRACKER_PATH)
        ws = wb["Registro de Pruebas"]
    except Exception as e:
        print(f"   ⚠️  No se pudo abrir el tracker: {e}")
        return False

    extra = extra or {}

    # Buscar última fila con datos
    last_row = 3
    for row in ws.iter_rows(min_row=4, max_col=1):
        if row[0].value is not None:
            last_row = row[0].row
        else:
            break
    new_row = last_row + 1
    run_id  = last_row - 2

    # Métricas
    wr     = metrics.get("win_rate", 0)
    sharpe = metrics.get("sharpe_ratio", 0)
    dd     = metrics.get("max_drawdown_pct", 0)
    trades = metrics.get("total_trades", 0)
    pf     = metrics.get("profit_factor", 0)
    cap    = metrics.get("final_capital", 150)
    pnl    = metrics.get("net_pnl_usd", 0)
    exp    = metrics.get("expectancy_usd", 0)

    # Trades por mes
    try:
        from config import BARS
        months = max(BARS.get("H4", 2000) * 4 / (24 * 30), 1)
    except Exception:
        months = 20
    tpm = round(trades / months, 1)

    verdict = _verdict(sharpe, wr, trades)
    ema_str = f"{extra.get('ema_fast', '?')}/{extra.get('ema_slow', '?')}"

    row_vals = [
        run_id,
        datetime.now().strftime("%Y-%m-%d"),
        version,
        config_name,
        ema_str,
        extra.get("adx_min", 15),
        extra.get("session", "london_ny"),
        trades,
        wr / 100,
        sharpe,
        dd / 100,
        pf,
        cap,
        pnl,
        exp,
        tpm,
        verdict,
        notes or strategy_mode,
    ]
    fmts = [
        None, "YYYY-MM-DD", None, None, None, "0", None,
        "0", "0.0%", "0.000", "0.00%", "0.000",
        "$#,##0.00", "$#,##0.00;[Red]($#,##0.00)",
        "$#,##0.0000", "0.0", None, None,
    ]

    if "EXCELENTE" in verdict:
        bg = "F0FDF4"
    elif "BUENO" in verdict:
        bg = "EFF6FF"
    elif "NEGATIVO" in verdict:
        bg = "FFF1F2"
    else:
        bg = "FFFBEB"

    for col_idx, (val, fmt) in enumerate(zip(row_vals, fmts), start=1):
        cell = ws.cell(row=new_row, column=col_idx, value=val)
        cell.font      = _font(bold=(col_idx in [1, 8, 9, 10, 13, 14]))
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(
            horizontal="center" if col_idx not in [4, 17, 18] else "left",
            vertical="center",
        )
        cell.border = _border()
        if fmt:
            cell.number_format = fmt
        if col_idx == 14 and isinstance(val, (int, float)):
            cell.font = _font(bold=True, color="065F46" if val >= 0 else "991B1B")
        if col_idx == 10 and isinstance(val, (int, float)):
            if val >= 2:    cell.font = _font(bold=True, color="065F46")
            elif val >= 1:  cell.font = _font(bold=True, color="1E40AF")
            elif val < 0:   cell.font = _font(bold=True, color="991B1B")

    ws.row_dimensions[new_row].height = 18

    try:
        wb.save(TRACKER_PATH)
        print(f"   📊 Tracker → fila {new_row} | {verdict}")
        return True
    except PermissionError:
        print("   ⚠️  Cierra el Excel y vuelve a ejecutar.")
        return False
    except Exception as e:
        print(f"   ⚠️  Error al guardar tracker: {e}")
        return False


def save_best_run(configs: dict, version: str = "v4") -> None:
    """Guarda todas las configs con trades > 0 en el tracker."""
    try:
        from config import INDICATORS
        extra = {
            "ema_fast": INDICATORS.get("ema_fast", 10),
            "ema_slow": INDICATORS.get("ema_slow", 50),
            "adx_min" : INDICATORS.get("adx_min",  15),
            "session" : "london_ny",
        }
    except Exception:
        extra = {}

    saved = 0
    for name, (_, _, m) in configs.items():
        if m.get("total_trades", 0) == 0:
            continue
        ok = save_run(
            metrics       = m,
            config_name   = name.strip(),
            version       = version,
            notes         = f"Auto v5.4 {datetime.now().strftime('%H:%M')}",
            strategy_mode = name,
            extra         = extra,
        )
        if ok:
            saved += 1
    if saved:
        print(f"   ✅ {saved} resultados guardados en trading_tracker.xlsx")


def export_trades_csv(trades: pd.DataFrame, filename: str) -> str:
    """Exporta trades detallados a CSV para auditoria reproducible."""
    from config import RESULTS_DIR

    os.makedirs(RESULTS_DIR, exist_ok=True)
    path = os.path.join(RESULTS_DIR, filename)
    trades.to_csv(path, index=False)
    return path
